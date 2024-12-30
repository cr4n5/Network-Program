from HDU_Login import newjw_login
import requests
import pickle
import os
import re
import json
import openpyxl
from openpyxl.styles import Border, Side

# 获取总课程表
def get_course(session):
    kc_url="https://newjw.hdu.edu.cn/jwglxt/kbcx/xskbcx_cxXsgrkb.html?gnmkdm=N253508"

    xnm , xqm = get_xnm_xqm(session)
    if xnm is None or xqm is None:
        print("获取学年学期信息失败")
        return None
    
    # 学年学期名称转换
    xnm = xnm.split("-")[0]
    if xqm == "1":
        xqm = "3"
    elif xqm == "2":
        xqm = "12"

    data={
        "xnm":xnm,
        "xqm":xqm,
        "kzlx":"ck",
        "xsdm":"",
    }

    try:
        response=session.post(kc_url,data=data)
    except requests.RequestException as e:
        print(f"HTTP request error: {e}")
        return None
    
    all_data = json.loads(response.text)
    kb_list = all_data["kbList"]

    return kb_list
   
# 获取当前学年和学期信息
def get_xnm_xqm(session):
    xnm_xqm_url="https://newjw.hdu.edu.cn/jwglxt/xtgl/index_cxAreaFive.html?localeKey=zh_CN&gnmkdm=index"
    data={
        "localeKey":"zh_CN",
        "gnmkdm":"index",
    }

    try:
        response=session.post(xnm_xqm_url,data=data)
    except requests.RequestException as e:
        print(f"HTTP request error: {e}")
        return None, None
    
    # 响应内容包含学年和学期信息
    content = response.text
    
    # 使用正则表达式匹配学年和学期信息
    match = re.search(r'(\d{4}-\d{4})学年(\d)学期', content)
    if match:
        xnm = match.group(1)
        xqm = match.group(2)
        print(f"学年: {xnm}, 学期: {xqm}")
        return xnm, xqm
    else:
        print("未找到匹配的学年和学期信息")
        return None, None

# 保存session
def save_session(session):
    with open("session.pkl","wb") as f:
        pickle.dump(session,f)

# 加载session登录
def login_session():
    if not os.path.exists("session.pkl"):
        session=requests.Session()
        session=newjw_login.login(session)
        if session is None:
            print("登录失败")
            exit()

        # 保存session
        save_session(session)
        return session
    else:
        with open("session.pkl","rb") as f:
            session=pickle.load(f)
    
    # 检查是否有效
    response=session.get("https://newjw.hdu.edu.cn/jwglxt/xtgl/login_slogin.html")
    if "用户登录" in response.text:
        print("登录过期")
        session=newjw_login.login(session)
        if session is None:
            exit()

        # 保存session
        save_session(session)
    return session

# 处理课程表
def process_course(course):
    
    processed_course=[]

    for item in course:
        # courese_detail
        course_name=item["kcmc"]
        room=item["cdmc"]
        teacher=item["xm"]
        week=item["zcd"]

        courese_detail=f"{course_name}\n{room}\n{teacher}\n{week}"
        day=item["xqjmc"]
        time=item["jcs"]
        
        time=time.split("-")
        if int(time[0])>=1 and int(time[1])<=5:
            mor_aft="上午"
        elif int(time[0])>=6 and int(time[1])<=9:
            mor_aft="下午"
        else:
            mor_aft="晚上"

        processed_course.append({"day":day,"time":time,"mor_aft":mor_aft,"course_detail":courese_detail})

    return processed_course
            
# 将处理完的课程表输出为excel
def output_excel(processed_course):
    # 设置边框
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
    columns = ['时间段','节次','星期一','星期二','星期三','星期四','星期五','星期六','星期日']
    # 使用openpyxl创建一个新的excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课程表"
    # 设置列宽
    for i in range(1,3):
        ws.column_dimensions[chr(64+i)].width = 10
    for i in range(3,10):
        ws.column_dimensions[chr(64+i)].width = 20
    # 设置行高
    for i in range(1,13):
        ws.row_dimensions[i].height = 40
    # 设置表头
    for i in range(1,10):
        ws.cell(row=1,column=i,value=columns[i-1]).border = border
    # 设置时间段
    ws.cell(row=2,column=1,value='上午')
    ws.cell(row=7,column=1,value='下午')
    ws.cell(row=11,column=1,value='晚上')
    # 合并单元格
    ws.merge_cells('A2:A6')
    ws.merge_cells('A7:A10')
    ws.merge_cells('A11:A13')
     # 设置合并单元格的对齐方式，设置外边框
    for cell in ['A2', 'A7', 'A11']:
        ws[cell].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws[cell].border = border
    # 设置节次
    for i in range(1,13):
        ws.cell(row=i+1,column=2,value=i).border = border
    # 设置课程表
    for item in processed_course:
        day = item['day']   
        time = item['time']
        mor_aft = item['mor_aft']
        course_detail = item['course_detail']
        # 合并time单元格
        column=columns.index(day)+1
        row=int(time[0])+1
        ws.merge_cells(start_row=row,start_column=column,end_row=int(time[1])+1,end_column=column)
        ws.cell(row=row,column=column,value=course_detail)
         # 设置合并单元格的对齐方式和自动换行
        ws.cell(row=row,column=column).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        for i in range(int(time[0])+1,int(time[1])+2):
            ws.cell(row=i,column=column).border = border

    # 将A1:I13的单元格设置外边框
    for row in ws.iter_rows(min_row=1, max_row=13, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
    # 保存文件
    wb.save("course.xlsx")
    print("课程表已保存为course.xlsx")
    

if __name__ == '__main__':
    
    # 登录
    session = login_session()

    # 获取课程表
    course=get_course(session)

    # 处理课程表
    processed_course=process_course(course)

    # 输出课程表
    output_excel(processed_course)